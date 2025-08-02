"""
FDOT City Data Explorer - Refactored with MVC Architecture
Main application entry point using the new MVC structure
"""

import streamlit as st
import logging

# Import MVC components
from models import CityCollection
from controllers import CityController
from views import MapView, CityView
from utils import load_css, create_header, create_footer
from utils.constants import UI_CONFIG

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class FDOTCityExplorer:
    """
    Main application class using MVC architecture
    """
    
    def __init__(self):
        """Initialize the application with MVC components"""
        self.city_controller = CityController()
        self.map_view = MapView()
        self.city_view = CityView()
        
        self._initialize_session_state()
    
    def _initialize_session_state(self):
        """Initialize session state variables"""
        if 'cities_data' not in st.session_state:
            st.session_state.cities_data = None
        if 'selected_city' not in st.session_state:
            st.session_state.selected_city = None
    
    def configure_page(self):
        """Configure Streamlit page settings"""
        st.set_page_config(
            page_title=UI_CONFIG['page_title'],
            page_icon=UI_CONFIG['page_icon'],
            layout=UI_CONFIG['layout'],
            initial_sidebar_state=UI_CONFIG['sidebar_state']
        )
    
    def render_header(self):
        """Render the simplified application header"""
        # Simple title following the wireframe design
        st.markdown(
            """
            <div style="text-align: center; padding: 1rem 0;">
                <h1 style="margin: 0; color: #1f77b4;">🗺️ FDOT City Data Explorer</h1>
            </div>
            """, 
            unsafe_allow_html=True
        )
    
    def render_sidebar(self) -> tuple:
        """
        Render the sidebar and handle data fetching
        
        Returns:
            Tuple of (action, params)
        """
        try:
            # Create smart sidebar controls
            action, params = self.city_view.create_smart_sidebar()
            
            # Handle data fetching
            if any(params.get('button', False) for key in ['button'] if key in params):
                success = self.city_view.handle_data_fetch(action, params)
                if success:
                    logger.info(f"Successfully executed action: {action}")
            
            # Display city data in sidebar if available
            cities = self.city_controller.get_session_cities()
            if cities:
                self.city_view.display_city_data_sidebar(cities)
            
            return action, params
            
        except Exception as e:
            logger.error(f"Error rendering sidebar: {e}")
            return "🌍 Fetch All Cities", {"limit": 50, "button": False}
    
    def render_main_map(self):
        """Render the main interactive map with new simplified layout"""
        try:
            # Get cities from session
            cities = self.city_controller.get_session_cities()
            
            if cities and len(cities) > 0:
                # Display map with new simplified UI
                self.map_view.display_cities_on_map(cities)
            else:
                # Show a simplified message and Florida map when no cities are loaded
                st.info("👆 Use the sidebar to fetch city data and start exploring!")
                self.map_view.display_florida_only_map()
                
        except Exception as e:
            logger.error(f"Error rendering main map: {e}")
            st.error("❌ Error displaying map. Please try refreshing the page.")
    
    def render_data_tabs(self):
        """Render the simplified data analysis tabs"""
        try:
            cities = self.city_controller.get_session_cities()
            
            if not cities or len(cities) == 0:
                return
            
            st.markdown("---")
            
            # Create simplified tabs for core functionality
            tab1, tab2, tab3, tab4 = st.tabs([
                "📊 City Data", 
                "📈 Analytics",
                "🚦 Traffic Data",
                "💾 Excel Export"
            ])
            
            with tab1:
                self.render_city_data_tab(cities)
            
            with tab2:
                self.render_analytics_tab(cities)
            
            with tab3:
                self.render_traffic_data_tab()
            
            with tab4:
                self.render_excel_export_tab(cities)
                
        except Exception as e:
            logger.error(f"Error rendering data tabs: {e}")
            st.error("❌ Error displaying data tabs")
    

    
    def render_city_data_tab(self, cities: CityCollection):
        """Render the city data tab"""
        try:
            st.markdown("### 📊 City Data Table")
            
            # Create filter controls
            filters = self.city_view.create_filter_controls(cities)
            
            # Display filtered city data
            self.city_view.display_city_data_main(cities, filters)
            
        except Exception as e:
            logger.error(f"Error rendering city data tab: {e}")
            st.error("❌ Error displaying city data")
    
    def render_analytics_tab(self, cities: CityCollection):
        """Render the analytics tab"""
        try:
            st.markdown("### 📈 City Analytics")
            self.city_view.create_charts(cities)
            
        except Exception as e:
            logger.error(f"Error rendering analytics tab: {e}")
            st.error("❌ Error displaying analytics")
    
    def render_traffic_data_tab(self):
        """Render the traffic data tab"""
        try:
            # Always try to get traffic data (from session or load fresh)
            traffic_data = st.session_state.get('traffic_data')
            if not traffic_data:
                # Try to load from files or fetch fresh
                traffic_data = self.city_controller.fetch_traffic_data()
                if traffic_data:
                    self.city_controller.save_traffic_data_to_json(traffic_data)
                    st.session_state.traffic_data = traffic_data
            
            if traffic_data:
                self.city_view.display_traffic_data(traffic_data)
            else:
                st.info("🚦 Traffic data is being loaded automatically. If you don't see traffic data, please wait a moment or refresh the page.")
                
                # Provide option to fetch traffic data manually
                if st.button("🚦 Fetch Traffic Data Now", type="primary"):
                    with st.spinner("🚦 Fetching traffic data..."):
                        traffic_data = self.city_controller.fetch_traffic_data()
                        if traffic_data:
                            self.city_controller.save_traffic_data_to_json(traffic_data)
                            st.session_state.traffic_data = traffic_data
                            st.success("✅ Traffic data fetched successfully!")
                            st.rerun()
                        else:
                            st.error("❌ Failed to fetch traffic data")
            
        except Exception as e:
            logger.error(f"Error rendering traffic data tab: {e}")
            st.error("❌ Error displaying traffic data")
    
    def render_excel_export_tab(self, cities: CityCollection):
        """Render the Excel export tab"""
        try:
            st.markdown("### 💾 Excel Export Center")
            st.info("📊 Export your data in professional Excel format with enhanced formatting and multiple sheets.")
            
            # City Data Export
            if cities and len(cities) > 0:
                st.markdown("#### 🏙️ City Data Export")
                df_cities = cities.to_dataframe()
                self.city_view.create_standalone_excel_export(df_cities, "City Data", "cities")
            
            # Traffic Data Export
            traffic_data = st.session_state.get('traffic_data')
            if traffic_data and 'features' in traffic_data:
                st.markdown("#### 🚦 Traffic Data Export")
                from models.city_model import TrafficDataCollection
                traffic_collection = TrafficDataCollection(traffic_data)
                df_traffic = traffic_collection.to_dataframe()
                self.city_view.create_standalone_excel_export(df_traffic, "Traffic Data", "traffic")
            
            # Combined Export Option
            if (cities and len(cities) > 0) and (traffic_data and 'features' in traffic_data):
                st.markdown("#### 🔗 Combined Data Export")
                st.info("💡 Export both city and traffic data in a single Excel file with multiple sheets.")
                
                if st.button("📊 Create Combined Excel Export", type="primary", use_container_width=True):
                    try:
                        import io
                        from openpyxl import Workbook
                        from openpyxl.utils.dataframe import dataframe_to_rows
                        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                        from openpyxl.utils import get_column_letter
                        
                        wb = Workbook()
                        
                        # City Data Sheet
                        ws_cities = wb.active
                        ws_cities.title = "City Data"
                        df_cities = cities.to_dataframe()
                        
                        # Add city data with styling
                        rows_cities = list(dataframe_to_rows(df_cities, index=False, header=True))
                        
                        # Style definitions
                        header_font = Font(bold=True, color="FFFFFF", size=12)
                        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                        header_alignment = Alignment(horizontal="center", vertical="center")
                        data_font = Font(size=10)
                        data_alignment = Alignment(horizontal="left", vertical="center")
                        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                     top=Side(style='thin'), bottom=Side(style='thin'))
                        
                        # Add city data
                        for col_idx, value in enumerate(rows_cities[0], 1):
                            cell = ws_cities.cell(row=1, column=col_idx, value=value)
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = header_alignment
                            cell.border = border
                        
                        for row_idx, row in enumerate(rows_cities[1:], 2):
                            for col_idx, value in enumerate(row, 1):
                                cell = ws_cities.cell(row=row_idx, column=col_idx, value=value)
                                cell.font = data_font
                                cell.alignment = data_alignment
                                cell.border = border
                        
                        # Traffic Data Sheet
                        ws_traffic = wb.create_sheet("Traffic Data")
                        traffic_collection = TrafficDataCollection(traffic_data)
                        df_traffic = traffic_collection.to_dataframe()
                        rows_traffic = list(dataframe_to_rows(df_traffic, index=False, header=True))
                        
                        # Add traffic data
                        for col_idx, value in enumerate(rows_traffic[0], 1):
                            cell = ws_traffic.cell(row=1, column=col_idx, value=value)
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = header_alignment
                            cell.border = border
                        
                        for row_idx, row in enumerate(rows_traffic[1:], 2):
                            for col_idx, value in enumerate(row, 1):
                                cell = ws_traffic.cell(row=row_idx, column=col_idx, value=value)
                                cell.font = data_font
                                cell.alignment = data_alignment
                                cell.border = border
                        
                        # Summary Sheet
                        ws_summary = wb.create_sheet("Summary")
                        ws_summary.cell(row=1, column=1, value="FDOT Data Export Summary").font = Font(bold=True, size=16)
                        ws_summary.cell(row=3, column=1, value=f"Export Date: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')}")
                        ws_summary.cell(row=4, column=1, value=f"City Records: {len(df_cities)}")
                        ws_summary.cell(row=5, column=1, value=f"Traffic Records: {len(df_traffic)}")
                        ws_summary.cell(row=6, column=1, value=f"Total Records: {len(df_cities) + len(df_traffic)}")
                        
                        # Auto-adjust column widths for all sheets
                        for ws in [ws_cities, ws_traffic]:
                            for column in ws.columns:
                                max_length = 0
                                column_letter = get_column_letter(column[0].column)
                                for cell in column:
                                    try:
                                        cell_length = len(str(cell.value)) if cell.value else 0
                                        if cell_length > max_length:
                                            max_length = cell_length
                                    except:
                                        pass
                                ws.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 50)
                        
                        # Freeze header rows
                        ws_cities.freeze_panes = "A2"
                        ws_traffic.freeze_panes = "A2"
                        
                        # Save to bytes
                        excel_buffer = io.BytesIO()
                        wb.save(excel_buffer)
                        excel_data = excel_buffer.getvalue()
                        
                        st.download_button(
                            label="📊 Download Combined Excel",
                            data=excel_data,
                            file_name=f"fdot_combined_data_{len(df_cities)}_{len(df_traffic)}_records.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            help="Download combined city and traffic data as Excel file",
                            use_container_width=True
                        )
                        
                        st.success("✅ Combined Excel file ready for download!")
                        
                    except Exception as e:
                        logger.error(f"Error creating combined Excel export: {e}")
                        st.error("❌ Combined Excel export failed")
            
            if not cities and not traffic_data:
                st.warning("⚠️ No data available for export. Please fetch city and/or traffic data first.")
                
        except Exception as e:
            logger.error(f"Error rendering Excel export tab: {e}")
            st.error("❌ Error displaying Excel export options")
    

    
    def render_simplified_data_tabs(self, cities: CityCollection):
        """
        Render simplified data tabs when Show Data button is pressed
        
        Args:
            cities: Collection of cities to display data for
        """
        try:
            # Create simplified tabs for data analysis
            tab1, tab2, tab3 = st.tabs([
                "📊 City Data", 
                "📈 Analytics",
                "🚦 Traffic Data"
            ])
            
            with tab1:
                # Simplified city data table
                filters = self.city_view.create_filter_controls(cities)
                self.city_view.display_city_data_main(cities, filters)
            
            with tab2:
                # Analytics
                self.city_view.create_charts(cities)
            
            with tab3:
                # Traffic data
                self.render_traffic_data_tab()
                
        except Exception as e:
            logger.error(f"Error rendering simplified data tabs: {e}")
            st.error("❌ Error displaying data analysis")
    
    def render_welcome_screen(self):
        """Render welcome screen when no data is available"""
        try:
            self.city_view.display_welcome_screen()
        except Exception as e:
            logger.error(f"Error rendering welcome screen: {e}")
    
    def render_footer(self):
        """Render the application footer"""
        create_footer()
    
    def run(self):
        """Run the main application with simplified workflow"""
        try:
            # Configure page
            self.configure_page()
            
            # Load custom CSS
            load_css()
            
            # Render header
            self.render_header()
            
            # Render sidebar and handle data fetching
            action, params = self.render_sidebar()
            
            # Render main map with integrated controls
            self.render_main_map()
            
            # Optional: Show data panel if requested (triggered by Show Data button)
            if st.session_state.get('show_data_panel', False):
                cities = self.city_controller.get_session_cities()
                if cities and len(cities) > 0:
                    st.markdown("---")
                    with st.expander("📊 Detailed Data Analysis", expanded=True):
                        self.render_simplified_data_tabs(cities)
            
            # Render footer
            self.render_footer()
            
        except Exception as e:
            logger.error(f"Critical error in main application: {e}")
            st.error("❌ Critical application error. Please refresh the page.")
            
            # Display error details in expander for debugging
            with st.expander("🔧 Error Details (for debugging)"):
                st.code(str(e))
                import traceback
                st.code(traceback.format_exc())


def main():
    """
    Main entry point for the application
    """
    try:
        app = FDOTCityExplorer()
        app.run()
    except Exception as e:
        logger.error(f"Failed to start application: {e}")
        st.error("❌ Failed to start application. Please check the configuration and try again.")


if __name__ == "__main__":
    main()